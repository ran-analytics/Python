#horizontal writing of xml to excel
import openpyxl

# Sample tag_value_dict for demonstration
tag_value_dict = {
    "root.child1": ["value1", "value2"],
    "root.child2": ["value3", "value4"],
    "root.child3": ["value5", "value6"]
}

wb = openpyxl.Workbook()
ws = wb.active

# Write headers (tag paths)
col = 1
for path in tag_value_dict:
    ws.cell(row=1, column=col).value = path
    col += 1

# Write text values horizontally (one row per tag path)
for row, (path, values) in enumerate(tag_value_dict.items(), start=2):
    col = 1
    for value in values:
        ws.cell(row=row, column=col).value = value
        col += 1

wb.save('parsed_data_horizontal.xlsx')
